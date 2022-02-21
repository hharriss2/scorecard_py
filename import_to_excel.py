import openpyxl as xl;

filename  = "/Users/harris-jones/Documents/coding2021/exceltest/person_fruit.xlsx"

wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

filename1 ="/Users/harris-jones/Documents/coding2021/exceltest/book_to_import.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active


mr = ws1.max_row
mc = ws1.max_column
  
# copying the cell values from source 
# excel file to destination excel file
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i, column = j).value = c.value
  
# saving the destination excel file
wb2.save(str(filename1))